#!/usr/bin/env python3
"""Validate the Mystcrag V1.0 simplified bead catalog workbook.

Usage:
    python validate_mystcrag_excel.py
    python validate_mystcrag_excel.py path/to/workbook.xlsx
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("缺少 openpyxl。请先运行：python -m pip install openpyxl") from exc


HERE = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = HERE / "玄矶珠子商品库导入模板_V1.0_简化优化版.xlsx"
SCHEMA_PATH = HERE / "mystcrag_schema.json"
SOURCE_PATH = HERE.parent / "bead-catalog-template-20260723/玄矶珠子商品库导入模板.xlsx"
SOURCE_SHA256 = "a9664e9b2b82a480befc7c863e2bcde72899fd0162b70ce7b7558e26b28705ae"
FORMAL_SHEETS = ["珠子SKU", "库存批次", "图片资产"]
STATUS_SHEETS = ["水晶主数据", "珠子SKU", "配件SKU", "库存批次", "图片资产"]
REQUIRED_SHEETS = [
    "使用说明",
    "水晶主数据",
    "珠子SKU",
    "配件SKU",
    "库存批次",
    "图片资产",
    "AI分析结果",
    "字段字典",
    "_枚举",
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_header(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("*", "")
    text = re.sub(r"（保留业务字段）$", "", text)
    return text


def header_map(ws) -> dict[str, int]:
    return {
        normalize_header(ws.cell(4, col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(4, col).value is not None
    }


def row_values(ws, row: int) -> dict[str, Any]:
    return {
        normalize_header(ws.cell(4, col).value): ws.cell(row, col).value
        for col in range(1, ws.max_column + 1)
        if ws.cell(4, col).value is not None
    }


def is_number(value: Any) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(float(value))
    )


def is_blank(value: Any) -> bool:
    return value is None or value == ""


def nonempty(values: Iterable[Any]) -> bool:
    return any(not is_blank(value) for value in values)


def check_enum(value: Any, allowed: list[str]) -> bool:
    return isinstance(value, str) and value in allowed


def bead_status(
    row: dict[str, Any],
    *,
    product_counts: Counter,
    sku_counts: Counter,
    crystal_ids: set[str],
    enums: dict[str, list[str]],
) -> str:
    business_values = [value for key, value in row.items() if key != "导入状态"]
    if not nonempty(business_values):
        return "待填写"

    dimension_names = ["直径_mm", "长度_mm", "宽度_mm", "高度_mm"]
    required = [
        "product_id",
        "sku",
        "crystal_id",
        "商品名称",
        "形状",
        "市场品质等级",
        "品质信息来源",
        "计价单位",
        "单位包含数量",
        "默认损耗率",
        "material_key",
        "商品主图文件名",
        "是否经过人工确认",
        "是否适合进入AI设计库",
        "是否上架",
        "币种",
        "售价_最小单位",
        "成本_最小单位",
        "可用库存",
    ]
    if any(is_blank(row.get(name)) for name in required) or all(
        is_blank(row.get(name)) for name in dimension_names
    ):
        return "请补全必填"
    if product_counts[row["product_id"]] > 1 or sku_counts[row["sku"]] > 1:
        return "ID重复"
    if row["crystal_id"] not in crystal_ids:
        return "关联不存在"

    enum_checks = [
        ("形状", "shape"),
        ("市场品质等级", "market_quality_grade"),
        ("品质信息来源", "quality_information_source"),
        ("计价单位", "pricing_unit"),
        ("是否经过人工确认", "boolean_text"),
        ("是否适合进入AI设计库", "boolean_text"),
        ("是否上架", "boolean_text"),
        ("币种", "currency"),
    ]
    if any(
        not check_enum(row.get(field), enums[enum_name])
        for field, enum_name in enum_checks
    ):
        return "枚举错误"
    if not is_blank(row.get("切面类型")) and not check_enum(
        row.get("切面类型"), enums["facet_type"]
    ):
        return "枚举错误"

    for name in ["直径_mm", "长度_mm", "宽度_mm", "高度_mm", "孔径_mm"]:
        value = row.get(name)
        if not is_blank(value) and (not is_number(value) or value <= 0):
            return "格式错误"
    for name in ["单位包含数量", "售价_最小单位", "成本_最小单位", "可用库存"]:
        value = row.get(name)
        if not is_number(value) or value < 0:
            return "格式错误"
    if row["单位包含数量"] <= 0:
        return "格式错误"
    for name in ["每串颗数", "串长_mm"]:
        value = row.get(name)
        if not is_blank(value) and (not is_number(value) or value <= 0):
            return "格式错误"
    rate = row.get("默认损耗率")
    if not is_number(rate) or rate < 0 or rate > 1:
        return "格式错误"
    display_order = row.get("展示排序")
    if not is_blank(display_order) and not is_number(display_order):
        return "格式错误"
    if row["是否适合进入AI设计库"] == "TRUE" and (
        row["是否经过人工确认"] != "TRUE"
        or row["市场品质等级"] == "UNKNOWN"
    ):
        return "待人工复核"
    return "可导入"


def lot_status(
    row: dict[str, Any],
    *,
    lot_counts: Counter,
    skus: set[str],
    enums: dict[str, list[str]],
) -> str:
    business_values = [value for key, value in row.items() if key != "导入状态"]
    if not nonempty(business_values):
        return "待填写"
    required = [
        "lot_id",
        "sku",
        "批次数量",
        "当前可用数量",
        "锁定数量",
        "单位",
        "批次成本",
        "币种",
        "市场品质等级",
        "品质信息来源",
        "是否有实拍图片",
        "是否经过人工确认",
        "是否可用于 AI 推荐",
        "质检状态",
    ]
    if any(is_blank(row.get(name)) for name in required):
        return "请补全必填"
    if lot_counts[row["lot_id"]] > 1:
        return "ID重复"
    if row["sku"] not in skus:
        return "关联不存在"
    enum_checks = [
        ("单位", "inventory_unit"),
        ("币种", "currency"),
        ("市场品质等级", "market_quality_grade"),
        ("品质信息来源", "quality_information_source"),
        ("是否有实拍图片", "boolean_text"),
        ("是否经过人工确认", "boolean_text"),
        ("是否可用于 AI 推荐", "boolean_text"),
        ("质检状态", "quality_check_status"),
    ]
    if any(
        not check_enum(row.get(field), enums[enum_name])
        for field, enum_name in enum_checks
    ):
        return "枚举错误"
    for name in ["批次数量", "当前可用数量", "锁定数量", "批次成本"]:
        value = row.get(name)
        if not is_number(value) or value < 0:
            return "格式错误"
    if row["当前可用数量"] + row["锁定数量"] > row["批次数量"]:
        return "格式错误"
    if row["是否可用于 AI 推荐"] == "TRUE" and (
        row["是否有实拍图片"] != "TRUE"
        or row["是否经过人工确认"] != "TRUE"
        or row["市场品质等级"] == "UNKNOWN"
        or row["质检状态"] != "PASSED"
    ):
        return "待人工复核"
    return "可导入"


def image_status(
    row: dict[str, Any],
    *,
    image_counts: Counter,
    crystal_ids: set[str],
    skus: set[str],
    lot_ids: set[str],
    enums: dict[str, list[str]],
) -> str:
    business_values = [value for key, value in row.items() if key != "导入状态"]
    if not nonempty(business_values):
        return "待填写"
    required = [
        "image_id",
        "文件名",
        "相对路径",
        "image_type",
        "是否真实实拍",
        "来源类型",
        "使用权限",
        "是否允许 AI 训练",
        "是否允许商业使用",
        "是否允许公开展示",
        "市场品质等级",
        "品质信息来源",
        "是否代表当前实际货品",
        "是否过度修图",
        "是否疑似 AI 图片",
        "人工复核状态",
    ]
    if any(is_blank(row.get(name)) for name in required) or all(
        is_blank(row.get(name)) for name in ["crystal_id", "sku", "lot_id"]
    ):
        return "请补全必填"
    if row.get("来源类型") in {"SUPPLIER", "PLATFORM", "WEB"} and is_blank(
        row.get("来源链接")
    ):
        return "请补全必填"
    if image_counts[row["image_id"]] > 1:
        return "ID重复"
    if (
        (not is_blank(row.get("crystal_id")) and row["crystal_id"] not in crystal_ids)
        or (not is_blank(row.get("sku")) and row["sku"] not in skus)
        or (not is_blank(row.get("lot_id")) and row["lot_id"] not in lot_ids)
    ):
        return "关联不存在"
    enum_checks = [
        ("image_type", "image_type"),
        ("是否真实实拍", "boolean_text"),
        ("来源类型", "image_source_type"),
        ("使用权限", "usage_permission"),
        ("是否允许 AI 训练", "boolean_text"),
        ("是否允许商业使用", "boolean_text"),
        ("是否允许公开展示", "boolean_text"),
        ("市场品质等级", "market_quality_grade"),
        ("品质信息来源", "quality_information_source"),
        ("是否代表当前实际货品", "boolean_text"),
        ("是否过度修图", "boolean_text"),
        ("是否疑似 AI 图片", "boolean_text"),
        ("人工复核状态", "review_status"),
    ]
    if any(
        not check_enum(row.get(field), enums[enum_name])
        for field, enum_name in enum_checks
    ):
        return "枚举错误"
    file_hash = row.get("文件哈希")
    if not is_blank(file_hash) and not (32 <= len(str(file_hash)) <= 128):
        return "格式错误"
    clear_permissions = {"OWNED", "AUTHORIZED", "PUBLIC_DOMAIN"}
    if (
        row["人工复核状态"] != "PASSED"
        or row["使用权限"] == "UNKNOWN"
        or (
            row["是否允许 AI 训练"] == "TRUE"
            and row["使用权限"] not in clear_permissions
        )
        or (
            row["是否允许商业使用"] == "TRUE"
            and row["使用权限"] not in clear_permissions
        )
        or row["是否过度修图"] == "TRUE"
        or row["是否疑似 AI 图片"] == "TRUE"
        or (
            row["是否代表当前实际货品"] == "TRUE"
            and row["是否真实实拍"] != "TRUE"
        )
    ):
        return "待人工复核"
    return "可导入"


class Report:
    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.passes: list[str] = []

    def check(self, condition: bool, success: str, failure: str) -> None:
        if condition:
            self.passes.append(success)
        else:
            self.errors.append(failure)

    def warn(self, condition: bool, message: str) -> None:
        if not condition:
            self.warnings.append(message)


def validate(workbook_path: Path) -> Report:
    report = Report()
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    enums = schema["enums"]

    report.check(workbook_path.exists(), "工作簿文件存在", f"找不到工作簿：{workbook_path}")
    if not workbook_path.exists():
        return report

    report.check(
        workbook_path.resolve() != SOURCE_PATH.resolve(),
        "输出文件与原始模板路径不同",
        "输出路径与原始模板相同，存在覆盖风险",
    )
    if SOURCE_PATH.exists():
        report.check(
            sha256(SOURCE_PATH) == SOURCE_SHA256,
            "原始模板 SHA-256 未变化",
            "原始模板哈希已变化，请确认是否被覆盖",
        )
    else:
        report.warnings.append(f"未找到原始模板，无法核对哈希：{SOURCE_PATH}")

    try:
        wb = load_workbook(workbook_path, data_only=False)
    except Exception as exc:
        report.errors.append(f"工作簿无法打开：{exc}")
        return report

    report.check(
        all(name in wb.sheetnames for name in REQUIRED_SHEETS),
        "所有要求的工作表均存在",
        f"缺少工作表：{[name for name in REQUIRED_SHEETS if name not in wb.sheetnames]}",
    )
    if not all(name in wb.sheetnames for name in REQUIRED_SHEETS):
        wb.close()
        return report

    report.check(wb["_枚举"].sheet_state == "hidden", "枚举辅助表已隐藏", "枚举辅助表未隐藏")
    report.check(
        wb.calculation is not None
        and wb.calculation.calcMode == "auto"
        and wb.calculation.fullCalcOnLoad,
        "工作簿设置为打开时自动完整重算",
        "计算模式未设置为自动完整重算",
    )

    for name in STATUS_SHEETS:
        ws = wb[name]
        formula_cells = [ws.cell(row, 1).value for row in range(5, ws.max_row + 1)]
        report.check(
            all(isinstance(value, str) and value.startswith("=") for value in formula_cells),
            f"{name} 状态公式覆盖全部预留数据行",
            f"{name} 存在缺失的状态公式",
        )
        report.check(
            all("#REF!" not in value for value in formula_cells if isinstance(value, str)),
            f"{name} 状态公式无 #REF!",
            f"{name} 状态公式包含 #REF!",
        )
        formula_text = "\n".join(value for value in formula_cells if isinstance(value, str))
        expected_formula_tokens = (
            ["COUNTIF", "LEN"] if name == "图片资产" else ["COUNTIF", "ISNUMBER"]
        )
        report.check(
            all(token in formula_text for token in expected_formula_tokens),
            f"{name} 包含重复/格式校验",
            f"{name} 缺少重复或格式校验",
        )
        report.check(
            len(ws.conditional_formatting) > 0,
            f"{name} 保留或扩展了条件格式",
            f"{name} 没有条件格式",
        )

    expected_headers = {
        sheet: [field["excel_header"] for field in details.get("fields", [])]
        for sheet, details in schema["sheets"].items()
        if details.get("fields")
    }
    for sheet, headers in expected_headers.items():
        actual = [wb[sheet].cell(4, col).value for col in range(1, len(headers) + 1)]
        report.check(
            actual == headers,
            f"{sheet} 表头与 schema 一致",
            f"{sheet} 表头与 schema 不一致",
        )

    prohibited = set(schema["prohibited_formal_entry_fields"])
    for sheet in FORMAL_SHEETS:
        headers = {
            normalize_header(wb[sheet].cell(4, col).value)
            for col in range(1, wb[sheet].max_column + 1)
        }
        found = sorted(prohibited & headers)
        report.check(
            not found,
            f"{sheet} 不含专业鉴定人工录入字段",
            f"{sheet} 仍含专业鉴定字段：{found}",
        )

    ai_schema = schema["sheets"]["AI分析结果"]
    report.check(
        ai_schema["all_fields_optional"]
        and not ai_schema["is_formal_import_source"]
        and not ai_schema["can_overwrite_other_sheets"],
        "AI 分析 schema 明确为全可选、非正式导入、不可覆盖",
        "AI 分析隔离规则不完整",
    )
    formal_formulas = "\n".join(
        str(cell.value)
        for sheet in STATUS_SHEETS
        for row in wb[sheet].iter_rows()
        for cell in row
        if cell.data_type == "f"
    )
    report.check(
        "AI分析结果" not in formal_formulas,
        "正式状态公式不引用 AI 分析结果",
        "正式状态公式错误引用 AI 分析结果",
    )

    report.check(
        len(wb["珠子SKU"].data_validations.dataValidation) >= 10,
        "珠子SKU 下拉和数字验证已配置",
        "珠子SKU 数据验证数量不足",
    )
    report.check(
        len(wb["库存批次"].data_validations.dataValidation) >= 8,
        "库存批次下拉和数字验证已配置",
        "库存批次数据验证数量不足",
    )
    report.check(
        len(wb["图片资产"].data_validations.dataValidation) >= 12,
        "图片资产下拉验证已配置",
        "图片资产数据验证数量不足",
    )
    report.check(
        len(wb["AI分析结果"].data_validations.dataValidation) >= 5,
        "AI 分析结果可选下拉和置信度验证已配置",
        "AI 分析结果数据验证数量不足",
    )

    crystal_ws = wb["水晶主数据"]
    bead_ws = wb["珠子SKU"]
    lot_ws = wb["库存批次"]
    image_ws = wb["图片资产"]
    crystal_ids = {
        crystal_ws.cell(row, 2).value
        for row in range(5, crystal_ws.max_row + 1)
        if not is_blank(crystal_ws.cell(row, 2).value)
    }
    bead_rows = [
        row_values(bead_ws, row)
        for row in range(5, bead_ws.max_row + 1)
        if nonempty(
            bead_ws.cell(row, col).value for col in range(2, bead_ws.max_column + 1)
        )
    ]
    product_counts = Counter(row["product_id"] for row in bead_rows)
    sku_counts = Counter(row["sku"] for row in bead_rows)
    skus = set(sku_counts)

    lot_rows = [
        row_values(lot_ws, row)
        for row in range(5, lot_ws.max_row + 1)
        if nonempty(lot_ws.cell(row, col).value for col in range(2, lot_ws.max_column + 1))
    ]
    lot_counts = Counter(row["lot_id"] for row in lot_rows)
    lot_ids = set(lot_counts)

    image_rows = [
        row_values(image_ws, row)
        for row in range(5, image_ws.max_row + 1)
        if nonempty(
            image_ws.cell(row, col).value for col in range(2, image_ws.max_column + 1)
        )
    ]
    image_counts = Counter(row["image_id"] for row in image_rows)

    actual_bead_statuses = [
        bead_status(
            row,
            product_counts=product_counts,
            sku_counts=sku_counts,
            crystal_ids=crystal_ids,
            enums=enums,
        )
        for row in bead_rows
    ]
    report.check(
        all(status == "可导入" for status in actual_bead_statuses),
        "现有珠子 SKU 示例独立校验均为可导入",
        f"珠子 SKU 示例存在问题：{actual_bead_statuses}",
    )
    actual_lot_statuses = [
        lot_status(row, lot_counts=lot_counts, skus=skus, enums=enums)
        for row in lot_rows
    ]
    report.check(
        all(status == "可导入" for status in actual_lot_statuses),
        "现有库存批次示例独立校验均为可导入",
        f"库存批次示例存在问题：{actual_lot_statuses}",
    )
    actual_image_statuses = [
        image_status(
            row,
            image_counts=image_counts,
            crystal_ids=crystal_ids,
            skus=skus,
            lot_ids=lot_ids,
            enums=enums,
        )
        for row in image_rows
    ]
    report.check(
        all(status == "可导入" for status in actual_image_statuses),
        "现有图片资产示例独立校验均为可导入",
        f"图片资产示例存在问题：{actual_image_statuses}",
    )

    if bead_rows:
        base = dict(bead_rows[0])
        blank_price = dict(base)
        blank_price["售价_最小单位"] = None
        report.check(
            bead_status(
                blank_price,
                product_counts=product_counts,
                sku_counts=sku_counts,
                crystal_ids=crystal_ids,
                enums=enums,
            )
            == "请补全必填",
            "空白必填数字不会被误判为有效",
            "空白售价未被识别为缺少必填",
        )

        duplicated_products = Counter(product_counts)
        duplicated_products[base["product_id"]] += 1
        report.check(
            bead_status(
                base,
                product_counts=duplicated_products,
                sku_counts=sku_counts,
                crystal_ids=crystal_ids,
                enums=enums,
            )
            == "ID重复",
            "重复 product_id 可被识别",
            "重复 product_id 校验失败",
        )

        bad_fk = dict(base)
        bad_fk["crystal_id"] = "crystal-does-not-exist"
        report.check(
            bead_status(
                bad_fk,
                product_counts=product_counts,
                sku_counts=sku_counts,
                crystal_ids=crystal_ids,
                enums=enums,
            )
            == "关联不存在",
            "错误 crystal_id 外键可被识别",
            "crystal_id 外键校验失败",
        )

        unknown_quality = dict(base)
        unknown_quality["市场品质等级"] = "UNKNOWN"
        report.check(
            bead_status(
                unknown_quality,
                product_counts=product_counts,
                sku_counts=sku_counts,
                crystal_ids=crystal_ids,
                enums=enums,
            )
            == "待人工复核",
            "UNKNOWN 品质且申请进入 AI 库时会待人工复核",
            "UNKNOWN 品质的 AI 入库复核门槛失效",
        )

    report.check(
        {"product-aquamarine-round-8", "product-moonstone-round-6"}.issubset(
            product_counts
        ),
        "原模板珠子示例数据已保留",
        "原模板珠子示例数据缺失",
    )
    report.check(
        "product-spacer-silver-3"
        in {
            wb["配件SKU"].cell(row, 2).value
            for row in range(5, wb["配件SKU"].max_row + 1)
        },
        "原模板配件示例数据已保留",
        "原模板配件示例数据缺失",
    )

    wb.close()
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="待校验的 .xlsx 文件；默认校验脚本同目录的优化版工作簿",
    )
    args = parser.parse_args()
    report = validate(args.workbook.resolve())

    print(f"校验文件：{args.workbook.resolve()}")
    for message in report.passes:
        print(f"[PASS] {message}")
    for message in report.warnings:
        print(f"[WARN] {message}")
    for message in report.errors:
        print(f"[FAIL] {message}")
    print(
        f"\n汇总：{len(report.passes)} passed, "
        f"{len(report.warnings)} warnings, {len(report.errors)} failed"
    )
    return 1 if report.errors else 0


if __name__ == "__main__":
    sys.exit(main())
